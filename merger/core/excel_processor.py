"""
Excel 合并核心逻辑
从原始 merge_excel.py 提取并适配
"""
import sys
import json
import re
from pathlib import Path
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

try:
    import xlwt
    XLS_SUPPORT = True
except ImportError:
    XLS_SUPPORT = False


def read_sheet(path):
    """Read the first worksheet, returning header, data rows, images, and header index."""
    if isinstance(path, str):
        path = Path(path)
    
    # First pass: load cached values (data_only=True) so formula cells resolve to their results
    wb_values = load_workbook(path, data_only=True)
    try:
        ws_values = wb_values.worksheets[0]
        rows = []
        for r in ws_values.iter_rows():
            rows.append(tuple(cell.value for cell in r))
    finally:
        wb_values.close()

    if len(rows) < 1:
        raise ValueError(f"{path} does not contain enough rows to find a header.")

    # Find first non-empty row index
    first_non_empty = 0
    for i, r in enumerate(rows):
        if any(cell is not None and str(cell).strip() != "" for cell in r):
            first_non_empty = i
            break

    # Search within a small window (first_non_empty .. first_non_empty+5) for the row
    # with the most non-empty cells — we assume that's the header row.
    search_end = min(len(rows), first_non_empty + 6)
    best_idx = first_non_empty
    best_count = -1
    for i in range(first_non_empty, search_end):
        count = sum(1 for cell in rows[i] if cell is not None and str(cell).strip() != "")
        if count > best_count:
            best_count = count
            best_idx = i

    # Normalize header cells to stripped strings (empty string for blank cells)
    header = tuple((str(cell).strip() if cell is not None else "") for cell in rows[best_idx])

    # Data rows are all rows after the detected header row
    data_rows = []
    # Detect columns that likely contain images (e.g., header contains "图" or "照")
    image_columns = {
        idx for idx, col_name in enumerate(header)
        if col_name and any(token in col_name for token in ("图", "照", "image", "photo"))
    }
    for row in rows[best_idx + 1:]:
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            normalized_row = list(row)
            for col_idx in image_columns:
                if col_idx < len(normalized_row):
                    normalized_row[col_idx] = None
            data_rows.append(tuple(normalized_row))

    # Second pass: load original workbook (data_only=False) to extract embedded images
    images_info = []
    wb_images = load_workbook(path, data_only=False)
    try:
        ws_images = wb_images.worksheets[0]
        if hasattr(ws_images, '_images') and ws_images._images:
            from openpyxl.utils import coordinate_from_string, column_index_from_string

            for img in ws_images._images:
                try:
                    img_bytes = None
                    if hasattr(img, '_data'):
                        img_bytes = img._data()
                    elif hasattr(img, 'ref') and hasattr(img.ref, 'read'):
                        img_bytes = img.ref.read()

                    if not img_bytes:
                        continue

                    anchor_info = {
                        'row': None,
                        'col': None,
                        'width': img.width if hasattr(img, 'width') else 100,
                        'height': img.height if hasattr(img, 'height') else 100,
                    }

                    if hasattr(img, 'anchor'):
                        anchor = img.anchor
                        if hasattr(anchor, '_from') and anchor._from:
                            anchor_info['row'] = anchor._from.row
                            anchor_info['col'] = anchor._from.col
                        elif isinstance(anchor, str):
                            col_str, row_num = coordinate_from_string(anchor)
                            anchor_info['row'] = row_num - 1  # zero-based
                            anchor_info['col'] = column_index_from_string(col_str) - 1

                    if anchor_info['row'] is not None and anchor_info['col'] is not None:
                        images_info.append({
                            'image_data': img_bytes,
                            'anchor_info': anchor_info,
                        })
                except Exception as exc:
                    print(f"Warning: Failed to extract image: {exc}", file=sys.stderr)
    finally:
        wb_images.close()

    return header, data_rows, images_info, best_idx


def merge_excel_files(file_paths):
    """合并多个Excel文件"""
    combined_header = []  # list of column names (strings)
    merged_rows = []  # list of lists aligned to combined_header
    all_images = []  # list of (row_offset, images_info) tuples

    current_row = 1  # Start from row 1 (0-indexed, row 0 is header)

    for path in file_paths:
        header, rows, images_info, header_row_idx = read_sheet(path)

        # Add any new columns to the combined header
        new_columns = []
        for col in header:
            if col and col not in combined_header:
                combined_header.append(col)
                new_columns.append(col)

        # If combined header grew, extend existing merged rows with None placeholders
        if new_columns:
            for r in merged_rows:
                r.extend([None] * len(new_columns))

        # Build index map from this file's column indices to combined_header indices
        index_map = []
        for col in header:
            if col and col in combined_header:
                index_map.append(combined_header.index(col))
            else:
                index_map.append(None)

        # Map each data row into the combined header order
        for row in rows:
            mapped = [None] * len(combined_header)
            for i, val in enumerate(row):
                if i < len(index_map) and index_map[i] is not None:
                    mapped[index_map[i]] = val
            merged_rows.append(mapped)

        # Store images with adjusted row offset
        for img_info in images_info:
            anchor_info = img_info['anchor_info']
            img_row = anchor_info['row']
            img_col = anchor_info['col']

            if img_col is None or img_col >= len(index_map):
                continue

            combined_col_idx = index_map[img_col]
            if combined_col_idx is None:
                continue

            data_row_idx = img_row - header_row_idx - 1

            if data_row_idx < 0:
                continue

            if data_row_idx >= len(rows):
                continue

            merged_row = current_row + data_row_idx + 1

            merged_row_idx = merged_row - 2
            if 0 <= merged_row_idx < len(merged_rows):
                row_values = merged_rows[merged_row_idx]
                if combined_col_idx < len(row_values):
                    row_values[combined_col_idx] = None

            all_images.append({
                'merged_row': merged_row,
                'col_index': combined_col_idx,
                'width': anchor_info['width'],
                'height': anchor_info['height'],
                'image_data': img_info['image_data'],
            })

        current_row += len(rows)

    return combined_header, merged_rows, all_images


def apply_cell_operations(merged_rows, combined_header, operations):
    """Apply cell operations to modify cell values in place."""
    if not operations:
        return
    
    for op in operations:
        col_name = op.get('column')
        action = op.get('action')
        
        # Find column index
        col_idx = None
        for i, name in enumerate(combined_header):
            if name == col_name:
                col_idx = i
                break
        
        if col_idx is None:
            print(f"Warning: Column '{col_name}' not found, skipping operation", file=sys.stderr)
            continue
        
        # Apply operation to each row
        for row in merged_rows:
            if col_idx >= len(row) or row[col_idx] is None:
                continue
            
            val = str(row[col_idx])
            
            if action == 'add_prefix':
                row[col_idx] = op['value'] + val
            elif action == 'add_suffix':
                row[col_idx] = val + op['value']
            elif action == 'remove_prefix':
                if val.startswith(op['value']):
                    row[col_idx] = val[len(op['value']):]
            elif action == 'remove_suffix':
                if val.endswith(op['value']):
                    row[col_idx] = val[:-len(op['value'])]
            elif action == 'replace':
                row[col_idx] = val.replace(op['old_value'], op['new_value'])
            elif action == 'insert_at':
                pos = op['position']
                if 0 <= pos <= len(val):
                    row[col_idx] = val[:pos] + op['value'] + val[pos:]
            elif action == 'delete_at':
                pos = op['position']
                length = op.get('length', 1)
                if 0 <= pos < len(val):
                    row[col_idx] = val[:pos] + val[pos + length:]


def create_derived_column(merged_rows, combined_header, rules):
    """Create a new column based on rules extracting from an existing column."""
    if not rules:
        return
    
    source_col = rules.get('source_column')
    new_col = rules.get('new_column')
    mappings = rules.get('mappings', [])
    extraction = rules.get('extraction')
    
    # Find source column index
    source_idx = None
    for i, name in enumerate(combined_header):
        if name == source_col:
            source_idx = i
            break
    
    if source_idx is None:
        print(f"Warning: Source column '{source_col}' not found", file=sys.stderr)
        return
    
    # Add new column to header
    combined_header.append(new_col)
    new_idx = len(combined_header) - 1
    
    # Process each row
    for row in merged_rows:
        # Extend row if needed
        if len(row) < len(combined_header):
            row.extend([None] * (len(combined_header) - len(row)))
        
        if source_idx >= len(row) or row[source_idx] is None:
            row[new_idx] = None
            continue
        
        source_val = str(row[source_idx])
        
        # Apply extraction if specified
        if extraction:
            start = extraction.get('start', 0)
            end = extraction.get('end', len(source_val))
            # 修复索引问题:
            # Python切片是[start:end),end是不包含的
            # 如果用户输入是1-based(位置5,6表示第5和第6个字符)
            # 需要转换为0-based: start=5-1=4, end=6(不减1,因为切片exclusive)
            if start > 0 and extraction.get('one_indexed', False):
                start -= 1
            # end不需要减1,因为Python切片的end是exclusive的
            # if end > 0 and extraction.get('one_indexed', False):
            #     end -= 1  # BUG: 这会导致少取一个字符
            extracted = source_val[start:end] if start < len(source_val) else ""
        else:
            extracted = source_val
        
        # 如果没有映射规则,直接使用提取的值
        result = extracted
        
        # Apply mappings (如果有)
        if mappings:
            for mapping in mappings:
                pattern = mapping.get('pattern')
                value = mapping.get('value')
                use_regex = mapping.get('regex', False)
                
                if use_regex:
                    if re.search(pattern, extracted):
                        result = value
                        break
                else:
                    # 修复映射匹配问题: 使用精确匹配而不是包含匹配
                    # 原来的 "pattern in extracted" 会导致 "1" 也能匹配 "01"
                    if pattern == extracted:
                        result = value
                        break
        
        row[new_idx] = result


def filter_columns(combined_header, merged_rows, filter_mode, filter_columns):
    """
    过滤列
    
    Args:
        combined_header: 表头列表
        merged_rows: 数据行列表
        filter_mode: 'keep' (只保留指定列) 或 'remove' (删除指定列)
        filter_columns: 要过滤的列名列表
    
    Returns:
        (filtered_header, filtered_rows): 过滤后的表头和数据行
    """
    if filter_mode == 'none' or not filter_columns:
        return combined_header, merged_rows
    
    # 找出要保留的列索引
    if filter_mode == 'keep':
        # 只保留指定的列
        keep_indices = []
        for i, col in enumerate(combined_header):
            if col in filter_columns:
                keep_indices.append(i)
    elif filter_mode == 'remove':
        # 删除指定的列
        keep_indices = []
        for i, col in enumerate(combined_header):
            if col not in filter_columns:
                keep_indices.append(i)
    else:
        return combined_header, merged_rows
    
    # 过滤表头
    filtered_header = [combined_header[i] for i in keep_indices]
    
    # 过滤数据行
    filtered_rows = []
    for row in merged_rows:
        filtered_row = [row[i] if i < len(row) else None for i in keep_indices]
        filtered_rows.append(filtered_row)
    
    return filtered_header, filtered_rows



def write_merged_excel(combined_header, merged_rows, output_path, all_images=None, output_format='xlsx'):
    """Write merged header and rows to an .xlsx or .xls file."""
    if isinstance(output_path, str):
        output_path = Path(output_path)
    
    # Determine format from extension or parameter
    if output_format == 'xls' or output_path.suffix.lower() == '.xls':
        if not XLS_SUPPORT:
            raise RuntimeError("XLS format requires xlwt library. Please install: pip install xlwt")
        
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Merged')
        
        if not combined_header:
            raise RuntimeError("No input data provided.")
        
        # Write header
        for col_idx, header_val in enumerate(combined_header):
            ws.write(0, col_idx, header_val)
        
        # Write data rows
        for row_idx, row in enumerate(merged_rows, start=1):
            for col_idx, cell_val in enumerate(row):
                if cell_val is not None:
                    ws.write(row_idx, col_idx, cell_val)
        
        # Note: xlwt doesn't support images
        if all_images:
            print("Warning: Image embedding not supported in XLS format. Images will be omitted.", file=sys.stderr)
        
        wb.save(str(output_path))
        return output_path
    
    # Write to XLSX format (default)
    wb = Workbook()
    ws = wb.active
    ws.title = "Merged"

    if not combined_header:
        raise RuntimeError("No input data provided.")

    ws.append(combined_header)
    for row in merged_rows:
        if len(row) < len(combined_header):
            row = list(row) + [None] * (len(combined_header) - len(row))
        ws.append(row)

    # Add images if provided
    if all_images:
        from openpyxl.utils import get_column_letter
        for img_data in all_images:
            try:
                img = OpenpyxlImage(BytesIO(img_data['image_data']))
                
                if img_data.get('width'):
                    img.width = img_data['width']
                if img_data.get('height'):
                    img.height = img_data['height']
                
                merged_row = img_data['merged_row']
                col_idx = img_data['col_index']
                
                cell_ref = f"{get_column_letter(col_idx + 1)}{merged_row}"
                img.anchor = cell_ref
                
                ws.add_image(img)
            except Exception as e:
                print(f"Warning: Failed to add image: {e}", file=sys.stderr)

    wb.save(str(output_path))
    return output_path


def filter_columns(combined_header, merged_rows, filter_mode, filter_columns):
    """
    根据过滤模式和列名列表过滤列
    
    Args:
        combined_header: 表头列表
        merged_rows: 数据行列表
        filter_mode: 过滤模式 ('none', 'keep', 'remove')
        filter_columns: 要过滤的列名列表
    
    Returns:
        filtered_header, filtered_rows: 过滤后的表头和数据
    """
    if filter_mode == 'none' or not filter_columns:
        return combined_header, merged_rows
    
    # 找到要保留的列索引
    if filter_mode == 'keep':
        # 只保留指定的列
        keep_indices = []
        filtered_header = []
        for i, col in enumerate(combined_header):
            if col in filter_columns:
                keep_indices.append(i)
                filtered_header.append(col)
    elif filter_mode == 'remove':
        # 删除指定的列
        keep_indices = []
        filtered_header = []
        for i, col in enumerate(combined_header):
            if col not in filter_columns:
                keep_indices.append(i)
                filtered_header.append(col)
    else:
        return combined_header, merged_rows
    
    # 过滤数据行
    filtered_rows = []
    for row in merged_rows:
        filtered_row = [row[i] if i < len(row) else None for i in keep_indices]
        filtered_rows.append(filtered_row)
    
    return filtered_header, filtered_rows
