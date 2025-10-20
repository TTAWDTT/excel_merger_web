"""
通用数据处理器
支持 Excel (XLS/XLSX)、CSV、JSON 格式的合并与处理
"""
import sys
import json
import csv
import re
from pathlib import Path
from io import StringIO, BytesIO
from typing import List, Tuple, Dict, Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

try:
    import xlwt
    XLS_SUPPORT = True
except ImportError:
    XLS_SUPPORT = False


class DataProcessor:
    """通用数据处理器类"""
    
    SUPPORTED_FORMATS = ['xlsx', 'xls', 'csv', 'json']
    
    @staticmethod
    def detect_format(file_path):
        """检测文件格式"""
        if isinstance(file_path, str):
            file_path = Path(file_path)
        
        suffix = file_path.suffix.lower()
        if suffix in ['.xlsx']:
            return 'xlsx'
        elif suffix in ['.xls']:
            return 'xls'
        elif suffix in ['.csv']:
            return 'csv'
        elif suffix in ['.json']:
            return 'json'
        else:
            raise ValueError(f"Unsupported file format: {suffix}")
    
    @staticmethod
    def read_file(file_path):
        """
        读取文件并返回标准化数据格式
        
        Returns:
            (header, data_rows, metadata): 表头、数据行、元数据
            - header: list of str
            - data_rows: list of list
            - metadata: dict (包含图片等额外信息)
        """
        file_format = DataProcessor.detect_format(file_path)
        
        if file_format == 'xlsx':
            return DataProcessor._read_excel(file_path, 'xlsx')
        elif file_format == 'xls':
            return DataProcessor._read_excel(file_path, 'xls')
        elif file_format == 'csv':
            return DataProcessor._read_csv(file_path)
        elif file_format == 'json':
            return DataProcessor._read_json(file_path)
        else:
            raise ValueError(f"Unsupported format: {file_format}")
    
    @staticmethod
    def _read_excel(path, format_type='xlsx'):
        """读取Excel文件 (从原 excel_processor.py 适配)"""
        if isinstance(path, str):
            path = Path(path)
        
        # First pass: load cached values
        wb_values = load_workbook(path, data_only=True)
        try:
            ws_values = wb_values.worksheets[0]
            rows = []
            for r in ws_values.iter_rows():
                rows.append(tuple(cell.value for cell in r))
        finally:
            wb_values.close()

        if len(rows) < 1:
            raise ValueError(f"{path} does not contain enough rows to find a header.")

        # Find first non-empty row index
        first_non_empty = 0
        for i, r in enumerate(rows):
            if any(cell is not None and str(cell).strip() != "" for cell in r):
                first_non_empty = i
                break

        # Search for header row
        search_end = min(len(rows), first_non_empty + 6)
        best_idx = first_non_empty
        best_count = -1
        for i in range(first_non_empty, search_end):
            count = sum(1 for cell in rows[i] if cell is not None and str(cell).strip() != "")
            if count > best_count:
                best_count = count
                best_idx = i

        # Normalize header
        header = tuple((str(cell).strip() if cell is not None else "") for cell in rows[best_idx])

        # Data rows
        data_rows = []
        image_columns = {
            idx for idx, col_name in enumerate(header)
            if col_name and any(token in col_name for token in ("图", "照", "image", "photo"))
        }
        
        for row in rows[best_idx + 1:]:
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                normalized_row = list(row)
                for col_idx in image_columns:
                    if col_idx < len(normalized_row):
                        normalized_row[col_idx] = None
                data_rows.append(normalized_row)

        # Extract images (only for xlsx)
        images_info = []
        if format_type == 'xlsx':
            wb_images = load_workbook(path, data_only=False)
            try:
                ws_images = wb_images.worksheets[0]
                if hasattr(ws_images, '_images') and ws_images._images:
                    from openpyxl.utils import coordinate_from_string, column_index_from_string

                    for img in ws_images._images:
                        try:
                            img_bytes = None
                            if hasattr(img, '_data'):
                                img_bytes = img._data()
                            elif hasattr(img, 'ref') and hasattr(img.ref, 'read'):
                                img_bytes = img.ref.read()

                            if not img_bytes:
                                continue

                            anchor_info = {
                                'row': None,
                                'col': None,
                                'width': img.width if hasattr(img, 'width') else 100,
                                'height': img.height if hasattr(img, 'height') else 100,
                            }

                            if hasattr(img, 'anchor'):
                                anchor = img.anchor
                                if hasattr(anchor, '_from') and anchor._from:
                                    anchor_info['row'] = anchor._from.row
                                    anchor_info['col'] = anchor._from.col
                                elif isinstance(anchor, str):
                                    col_str, row_num = coordinate_from_string(anchor)
                                    anchor_info['row'] = row_num - 1
                                    anchor_info['col'] = column_index_from_string(col_str) - 1

                            if anchor_info['row'] is not None and anchor_info['col'] is not None:
                                images_info.append({
                                    'image_data': img_bytes,
                                    'anchor_info': anchor_info,
                                    'header_row_idx': best_idx,
                                })
                        except Exception as exc:
                            print(f"Warning: Failed to extract image: {exc}", file=sys.stderr)
            finally:
                wb_images.close()

        metadata = {
            'images': images_info,
            'header_row_idx': best_idx,
            'format': format_type,
        }
        
        return list(header), data_rows, metadata
    
    @staticmethod
    def _read_csv(path, encoding='utf-8'):
        """读取CSV文件"""
        if isinstance(path, str):
            path = Path(path)
        
        # Try different encodings
        encodings = [encoding, 'utf-8-sig', 'gbk', 'gb2312', 'latin1']
        
        for enc in encodings:
            try:
                with open(path, 'r', encoding=enc, newline='') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    
                    if not rows:
                        raise ValueError(f"{path} is empty")
                    
                    # First non-empty row is header
                    header = None
                    data_start = 0
                    
                    for i, row in enumerate(rows):
                        if any(cell.strip() for cell in row):
                            header = [cell.strip() for cell in row]
                            data_start = i + 1
                            break
                    
                    if header is None:
                        raise ValueError(f"{path} has no valid header")
                    
                    # Data rows
                    data_rows = []
                    for row in rows[data_start:]:
                        if any(cell.strip() for cell in row):
                            # Normalize row length to match header
                            normalized_row = row[:len(header)]
                            if len(normalized_row) < len(header):
                                normalized_row.extend([''] * (len(header) - len(normalized_row)))
                            data_rows.append(normalized_row)
                    
                    metadata = {
                        'format': 'csv',
                        'encoding': enc,
                    }
                    
                    return header, data_rows, metadata
                    
            except (UnicodeDecodeError, Exception) as e:
                if enc == encodings[-1]:
                    raise ValueError(f"Failed to read CSV file with any encoding: {e}")
                continue
    
    @staticmethod
    def _read_json(path, encoding='utf-8'):
        """
        读取JSON文件
        支持格式:
        1. Array of objects: [{"col1": "val1", "col2": "val2"}, ...]
        2. Object with data array: {"data": [...]}
        3. Object with records: {"records": [...]}
        """
        if isinstance(path, str):
            path = Path(path)
        
        with open(path, 'r', encoding=encoding) as f:
            data = json.load(f)
        
        # Determine structure
        records = None
        
        if isinstance(data, list):
            records = data
        elif isinstance(data, dict):
            # Try common keys
            for key in ['data', 'records', 'rows', 'items']:
                if key in data and isinstance(data[key], list):
                    records = data[key]
                    break
            
            if records is None:
                raise ValueError(f"JSON structure not recognized. Expected array or object with 'data'/'records' key")
        else:
            raise ValueError(f"JSON must be an array or object")
        
        if not records:
            raise ValueError(f"{path} contains no data")
        
        # Extract header from first record
        if not isinstance(records[0], dict):
            raise ValueError(f"JSON records must be objects/dictionaries")
        
        header = list(records[0].keys())
        
        # Extract data rows
        data_rows = []
        for record in records:
            row = [record.get(col, None) for col in header]
            # Convert to string for consistency
            row = [str(val) if val is not None else '' for val in row]
            data_rows.append(row)
        
        metadata = {
            'format': 'json',
            'encoding': encoding,
        }
        
        return header, data_rows, metadata
    
    @staticmethod
    def merge_files(file_paths, output_format='xlsx'):
        """
        合并多个文件
        
        Args:
            file_paths: 文件路径列表
            output_format: 输出格式
        
        Returns:
            (header, merged_rows, all_metadata): 合并后的表头、数据、元数据
        """
        combined_header = []
        merged_rows = []
        all_images = []
        current_row = 1
        
        for path in file_paths:
            header, rows, metadata = DataProcessor.read_file(path)
            
            # Add new columns to combined header
            new_columns = []
            for col in header:
                if col and col not in combined_header:
                    combined_header.append(col)
                    new_columns.append(col)
            
            # Extend existing rows if header grew
            if new_columns:
                for r in merged_rows:
                    r.extend([None] * len(new_columns))
            
            # Build index map
            index_map = []
            for col in header:
                if col and col in combined_header:
                    index_map.append(combined_header.index(col))
                else:
                    index_map.append(None)
            
            # Map rows to combined header
            for row in rows:
                mapped = [None] * len(combined_header)
                for i, val in enumerate(row):
                    if i < len(index_map) and index_map[i] is not None:
                        mapped[index_map[i]] = val
                merged_rows.append(mapped)
            
            # Process images (only for Excel files)
            if metadata.get('format') in ['xlsx', 'xls'] and metadata.get('images'):
                header_row_idx = metadata.get('header_row_idx', 0)
                
                for img_info in metadata['images']:
                    anchor_info = img_info['anchor_info']
                    img_row = anchor_info['row']
                    img_col = anchor_info['col']
                    
                    if img_col is None or img_col >= len(index_map):
                        continue
                    
                    combined_col_idx = index_map[img_col]
                    if combined_col_idx is None:
                        continue
                    
                    data_row_idx = img_row - header_row_idx - 1
                    if data_row_idx < 0 or data_row_idx >= len(rows):
                        continue
                    
                    merged_row = current_row + data_row_idx + 1
                    merged_row_idx = merged_row - 2
                    
                    if 0 <= merged_row_idx < len(merged_rows):
                        row_values = merged_rows[merged_row_idx]
                        if combined_col_idx < len(row_values):
                            row_values[combined_col_idx] = None
                    
                    all_images.append({
                        'merged_row': merged_row,
                        'col_index': combined_col_idx,
                        'width': anchor_info['width'],
                        'height': anchor_info['height'],
                        'image_data': img_info['image_data'],
                    })
            
            current_row += len(rows)
        
        all_metadata = {
            'images': all_images,
            'output_format': output_format,
        }
        
        return combined_header, merged_rows, all_metadata
    
    @staticmethod
    def apply_cell_operations(merged_rows, combined_header, operations):
        """应用单元格操作"""
        if not operations:
            return
        
        for op in operations:
            col_name = op.get('column')
            action = op.get('action')
            
            # Find column index
            col_idx = None
            for i, name in enumerate(combined_header):
                if name == col_name:
                    col_idx = i
                    break
            
            if col_idx is None:
                print(f"Warning: Column '{col_name}' not found", file=sys.stderr)
                continue
            
            # Apply operation to each row
            for row in merged_rows:
                if col_idx >= len(row) or row[col_idx] is None:
                    continue
                
                val = str(row[col_idx]) if row[col_idx] is not None else ''
                
                if action == 'add_prefix':
                    row[col_idx] = op['value'] + val
                elif action == 'add_suffix':
                    row[col_idx] = val + op['value']
                elif action == 'remove_prefix':
                    if val.startswith(op['value']):
                        row[col_idx] = val[len(op['value']):]
                elif action == 'remove_suffix':
                    if val.endswith(op['value']):
                        row[col_idx] = val[:-len(op['value'])]
                elif action == 'replace':
                    row[col_idx] = val.replace(op.get('old_value', ''), op.get('new_value', ''))
                elif action == 'insert_at':
                    pos = op.get('position', 0)
                    if 0 <= pos <= len(val):
                        row[col_idx] = val[:pos] + op['value'] + val[pos:]
                elif action == 'delete_at':
                    pos = op.get('position', 0)
                    length = op.get('length', 1)
                    if 0 <= pos < len(val):
                        row[col_idx] = val[:pos] + val[pos + length:]
    
    @staticmethod
    def create_derived_column(merged_rows, combined_header, rules):
        """创建派生列"""
        if not rules:
            return
        
        source_col = rules.get('source_column')
        new_col = rules.get('new_column')
        mappings = rules.get('mappings', [])
        extraction = rules.get('extraction')
        
        # Find source column index
        source_idx = None
        for i, name in enumerate(combined_header):
            if name == source_col:
                source_idx = i
                break
        
        if source_idx is None:
            print(f"Warning: Source column '{source_col}' not found", file=sys.stderr)
            return
        
        # Add new column
        combined_header.append(new_col)
        new_idx = len(combined_header) - 1
        
        # Process each row
        for row in merged_rows:
            # Extend row if needed
            if len(row) < len(combined_header):
                row.extend([None] * (len(combined_header) - len(row)))
            
            if source_idx >= len(row) or row[source_idx] is None:
                row[new_idx] = None
                continue
            
            source_val = str(row[source_idx])
            
            # Apply extraction
            if extraction:
                start = extraction.get('start', 0)
                end = extraction.get('end', len(source_val))
                if start > 0 and extraction.get('one_indexed', False):
                    start -= 1
                extracted = source_val[start:end] if start < len(source_val) else ""
            else:
                extracted = source_val
            
            result = extracted
            
            # Apply mappings
            if mappings:
                for mapping in mappings:
                    pattern = mapping.get('pattern')
                    value = mapping.get('value')
                    use_regex = mapping.get('regex', False)
                    
                    if use_regex:
                        if re.search(pattern, extracted):
                            result = value
                            break
                    else:
                        if pattern == extracted:
                            result = value
                            break
            
            row[new_idx] = result
    
    @staticmethod
    def filter_columns(combined_header, merged_rows, filter_mode, filter_columns):
        """过滤列"""
        if filter_mode == 'none' or not filter_columns:
            return combined_header, merged_rows
        
        # Determine columns to keep
        if filter_mode == 'keep':
            keep_indices = [i for i, col in enumerate(combined_header) if col in filter_columns]
        elif filter_mode == 'remove':
            keep_indices = [i for i, col in enumerate(combined_header) if col not in filter_columns]
        else:
            return combined_header, merged_rows
        
        # Filter header
        filtered_header = [combined_header[i] for i in keep_indices]
        
        # Filter rows
        filtered_rows = []
        for row in merged_rows:
            filtered_row = [row[i] if i < len(row) else None for i in keep_indices]
            filtered_rows.append(filtered_row)
        
        return filtered_header, filtered_rows
    
    @staticmethod
    def write_file(combined_header, merged_rows, output_path, metadata=None, output_format='xlsx'):
        """
        写入文件
        
        Args:
            combined_header: 表头
            merged_rows: 数据行
            output_path: 输出路径
            metadata: 元数据 (包含图片等)
            output_format: 输出格式
        """
        if isinstance(output_path, str):
            output_path = Path(output_path)
        
        # Override format from path if specified
        if output_path.suffix:
            detected_format = output_path.suffix.lower().lstrip('.')
            if detected_format in DataProcessor.SUPPORTED_FORMATS:
                output_format = detected_format
        
        if output_format == 'xlsx':
            return DataProcessor._write_xlsx(combined_header, merged_rows, output_path, metadata)
        elif output_format == 'xls':
            return DataProcessor._write_xls(combined_header, merged_rows, output_path)
        elif output_format == 'csv':
            return DataProcessor._write_csv(combined_header, merged_rows, output_path)
        elif output_format == 'json':
            return DataProcessor._write_json(combined_header, merged_rows, output_path)
        else:
            raise ValueError(f"Unsupported output format: {output_format}")
    
    @staticmethod
    def _write_xlsx(combined_header, merged_rows, output_path, metadata=None):
        """写入XLSX文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged"
        
        if not combined_header:
            raise RuntimeError("No header provided")
        
        # Write header
        ws.append(combined_header)
        
        # Write data rows
        for row in merged_rows:
            if len(row) < len(combined_header):
                row = list(row) + [None] * (len(combined_header) - len(row))
            ws.append(row)
        
        # Add images
        if metadata and metadata.get('images'):
            from openpyxl.utils import get_column_letter
            for img_data in metadata['images']:
                try:
                    img = OpenpyxlImage(BytesIO(img_data['image_data']))
                    
                    if img_data.get('width'):
                        img.width = img_data['width']
                    if img_data.get('height'):
                        img.height = img_data['height']
                    
                    merged_row = img_data['merged_row']
                    col_idx = img_data['col_index']
                    
                    cell_ref = f"{get_column_letter(col_idx + 1)}{merged_row}"
                    img.anchor = cell_ref
                    
                    ws.add_image(img)
                except Exception as e:
                    print(f"Warning: Failed to add image: {e}", file=sys.stderr)
        
        wb.save(str(output_path))
        return output_path
    
    @staticmethod
    def _write_xls(combined_header, merged_rows, output_path):
        """写入XLS文件"""
        if not XLS_SUPPORT:
            raise RuntimeError("XLS format requires xlwt library. Install: pip install xlwt")
        
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Merged')
        
        if not combined_header:
            raise RuntimeError("No header provided")
        
        # Write header
        for col_idx, header_val in enumerate(combined_header):
            ws.write(0, col_idx, header_val)
        
        # Write data rows
        for row_idx, row in enumerate(merged_rows, start=1):
            for col_idx, cell_val in enumerate(row):
                if cell_val is not None:
                    ws.write(row_idx, col_idx, cell_val)
        
        wb.save(str(output_path))
        return output_path
    
    @staticmethod
    def _write_csv(combined_header, merged_rows, output_path, encoding='utf-8-sig'):
        """写入CSV文件"""
        with open(output_path, 'w', encoding=encoding, newline='') as f:
            writer = csv.writer(f)
            
            # Write header
            writer.writerow(combined_header)
            
            # Write data rows
            for row in merged_rows:
                # Ensure row length matches header
                if len(row) < len(combined_header):
                    row = list(row) + [''] * (len(combined_header) - len(row))
                # Convert None to empty string
                row = [str(val) if val is not None else '' for val in row]
                writer.writerow(row)
        
        return output_path
    
    @staticmethod
    def _write_json(combined_header, merged_rows, output_path, encoding='utf-8'):
        """
        写入JSON文件
        格式: {"data": [{"col1": "val1", ...}, ...]}
        """
        records = []
        for row in merged_rows:
            record = {}
            for i, col in enumerate(combined_header):
                val = row[i] if i < len(row) else None
                # Convert None to empty string for JSON
                record[col] = val if val is not None else ''
            records.append(record)
        
        output_data = {
            'data': records,
            'count': len(records),
            'columns': combined_header,
        }
        
        with open(output_path, 'w', encoding=encoding, ensure_ascii=False) as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)
        
        return output_path
